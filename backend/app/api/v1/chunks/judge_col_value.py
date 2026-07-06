import numpy as np
from openpyxl import load_workbook
from typing import List, Optional, Tuple
import os
import asyncio
import sys

# 添加项目根目录到路径，以便导入模块
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "../../../../"))
from app.services.text_splitter.semantic_embedding import (
    create_embedding_provider,
    EmbeddingProvider
)

# ==============================
# 配置
# ==============================
DELTA_THRESHOLD = 0.1  # 组间相似度阈值
MAX_SCAN_ROWS = 10   # 最多扫描 10 个有效行
MAX_SCAN_COLS = 10   # 最多扫描 10 个有效列
MIN_COL_SAMPLES = 3

# 全局 embedding provider 实例（懒加载）
_embedding_provider_instance = None


def set_embedding_config(provider: str = None, api_key: str = None, base_url: str = None, model: str = None):
    """设置全局 embedding 配置并创建 provider 实例"""
    global _embedding_provider_instance
    if api_key:
        _embedding_provider_instance = create_embedding_provider(
            provider=provider or "openai",
            api_key=api_key,
            base_url=base_url,
            model=model
        )


def get_embedding_provider() -> EmbeddingProvider:
    """获取全局 embedding provider 实例（懒加载）"""
    global _embedding_provider_instance
    if _embedding_provider_instance is None:
        raise RuntimeError(
            "Embedding 模型未配置，请先在模型配置中添加 embedding 类型的模型并设为默认。"
            "调用 set_embedding_config() 或在系统设置中配置 embedding 模型。"
        )
    return _embedding_provider_instance


def clean_text(text: str) -> str:
    """清理文本：移除控制字符、多余空白，限制长度"""
    import re
    # 移除控制字符
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    # 合并多余空白
    text = re.sub(r'\s+', ' ', text).strip()
    # 限制长度（DashScope API 有长度限制）
    if len(text) > 8000:
        text = text[:8000]
    return text


async def get_embeddings_async(texts: List[str]) -> List[List[float]]:
    """异步获取 embeddings"""
    # 清理文本
    cleaned_texts = [clean_text(t) for t in texts]

    # 打印清理前后的对比
    if texts != cleaned_texts:
        print(f"[DEBUG] Texts cleaned: {len(texts)} -> {len(cleaned_texts)}")
        for i in range(min(3, len(texts))):
            if texts[i] != cleaned_texts[i]:
                print(f"[DEBUG] Text {i} cleaned: {repr(texts[i])[:50]} -> {repr(cleaned_texts[i])[:50]}")

    provider = get_embedding_provider()

    # DashScope API 限制批量大小不超过 10
    batch_size = 10
    all_embeddings = []

    try:
        # 分批处理
        for i in range(0, len(cleaned_texts), batch_size):
            batch = cleaned_texts[i:i + batch_size]
            batch_num = i // batch_size + 1
            total_batches = (len(cleaned_texts) + batch_size - 1) // batch_size

            print(f"[DEBUG] Calling provider.get_embeddings batch {batch_num}/{total_batches} with {len(batch)} texts")

            batch_result = await provider.get_embeddings(batch)
            all_embeddings.extend(batch_result)
            print(f"[DEBUG] Batch {batch_num}/{total_batches} completed, got {len(batch_result)} embeddings")

        print(f"[DEBUG] Successfully got {len(all_embeddings)} embeddings total")
        return all_embeddings
    except Exception as e:
        # 打印调试信息
        print(f"[DEBUG] Embedding API error: {e}")
        # 尝试获取响应体（如果是 HTTP 错误）
        try:
            import httpx
            if isinstance(e, httpx.HTTPStatusError) and hasattr(e, 'response'):
                print(f"[DEBUG] Response status: {e.response.status_code}")
                print(f"[DEBUG] Response body: {e.response.text[:500]}")
        except:
            pass
        print(f"[DEBUG] Number of texts: {len(cleaned_texts)}")
        # 打印前几个文本（限制长度）
        for i, text in enumerate(cleaned_texts[:5]):
            print(f"[DEBUG] Text {i} ({len(text)} chars): {repr(text)[:100]}")
        # 尝试打印更多错误信息
        import traceback
        print(f"[DEBUG] Traceback: {traceback.format_exc()}")
        raise


def get_embeddings_sync(texts: List[str]) -> List[List[float]]:
    """同步获取 embeddings（在同步函数中调用异步 API）"""
    try:
        # 尝试获取现有事件循环
        loop = asyncio.get_event_loop()
    except RuntimeError:
        # 没有事件循环，创建一个新的
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    # 运行异步函数
    return loop.run_until_complete(get_embeddings_async(texts))


def load_grid(file_path: str, sheet_name: Optional[str] = None) -> List[List]:
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return grid


def safe_str(value) -> str:
    return str(value).strip() if value is not None else ""


def cosine_similarity(a: np.ndarray, b: np.ndarray) -> float:
    norm_a, norm_b = np.linalg.norm(a), np.linalg.norm(b)
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return float(np.dot(a, b) / (norm_a * norm_b))


def compute_column_consistency(col_values: List[str]) -> float:
    texts = [t for t in col_values if t]
    if len(texts) < MIN_COL_SAMPLES:
        return 0.0

    # 使用同步 embedding 调用
    embeddings = get_embeddings_sync(texts)
    n = len(embeddings)
    total_sim = 0.0
    count = 0
    for i in range(n):
        for j in range(i + 1, n):
            sim = cosine_similarity(np.array(embeddings[i]), np.array(embeddings[j]))
            total_sim += sim
            count += 1
    return total_sim / count if count > 0 else 0.0


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
# 新增：单元格级行对相似度计算（正则化核心）
# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
def compute_row_pair_similarity_by_cells(
    grid: List[List],
    row_i: int,
    row_j: int,
    start_col: int,
    num_cols: int
) -> float:
    """
    计算两行在指定列范围内的单元格级平均 embedding 相似度。
    - 仅当两个单元格都非空时参与计算
    - 返回 [0, 1] 的平均相似度
    """
    sims = []
    texts_a = []
    texts_b = []
    positions = []

    for c in range(start_col, min(start_col + num_cols, len(grid[0]))):
        a = safe_str(grid[row_i][c])
        b = safe_str(grid[row_j][c])
        if a and b:  # 都非空才计算
            texts_a.append(a)
            texts_b.append(b)
            positions.append(c)

    if not texts_a:
        return 0.0

    # 批量获取 embeddings（更高效）
    all_texts = texts_a + texts_b
    embeddings = get_embeddings_sync(all_texts)

    half = len(texts_a)
    for idx in range(half):
        emb_a = np.array(embeddings[idx])
        emb_b = np.array(embeddings[half + idx])
        sim = cosine_similarity(emb_a, emb_b)
        sims.append(sim)

    return sum(sims) / len(sims) if sims else 0.0


def detect_table_boundary_skip_empty_start(grid):
    if not grid or not grid[0]:
        raise ValueError("表格为空！")

    rows, cols = len(grid), len(grid[0])
    print(f"📊 表格尺寸：{rows} 行 x {cols} 列")

    # --------------------------------------------------
    # 0. 跳过开头为空的行和列
    # --------------------------------------------------
    first_valid_row = 0
    for i in range(rows):
        if cols > 0 and safe_str(grid[i][0]) != "":
            first_valid_row = i
            break
    else:
        first_valid_row = 0

    first_valid_col = 0
    for j in range(cols):
        if rows > 0 and j < len(grid[0]) and safe_str(grid[0][j]) != "":
            first_valid_col = j
            break
    else:
        first_valid_col = 0

    print(f"⏭️ 跳过前导空白：从 (行={first_valid_row}, 列={first_valid_col}) 开始分析")

    effective_rows = rows - first_valid_row
    effective_cols = cols - first_valid_col
    if effective_rows <= 0 or effective_cols <= 0:
        print("⚠️ 有效区域为空，返回 (0, 0)")
        return 0, 0

    print("🧠 使用在线 embedding API...")

    # --------------------------------------------------
    # 1. 行边界检测（使用单元格级正则化相似度）
    # --------------------------------------------------
    print("\n➡️ 检测行边界（单元格级 embedding 相似度，最多 10 行）...")
    header_row_end = first_valid_row

    scan_rows = min(MAX_SCAN_ROWS, effective_rows)
    if scan_rows >= 2:
        row_sims = []
        scan_cols = min(MAX_SCAN_COLS, effective_cols)

        for i in range(scan_rows - 1):
            actual_i = first_valid_row + i
            actual_i1 = first_valid_row + i + 1
            sim = compute_row_pair_similarity_by_cells(
                grid, actual_i, actual_i1,
                start_col=first_valid_col,
                num_cols=scan_cols
            )
            row_sims.append(sim)
            print(f"  有效行 {i}-{i+1} (实际行 {actual_i}-{actual_i1}): 单元格级相似度 = {sim:.3f}")

        # 寻找相似度突增点
        for i in range(1, len(row_sims)):
            delta = row_sims[i] - row_sims[i - 1]
            print(f"    Δ(有效行{i}) = {delta:.3f}")
            if delta > DELTA_THRESHOLD:
                header_row_end = first_valid_row + i
                print(f"✅ 行边界确定：数据从实际行 {header_row_end} 开始")
                break
        else:
            print("ℹ️ 未检测到行表头（数据从 first_valid_row 开始）")
    else:
        print("ℹ️ 有效行数不足，数据从 first_valid_row 开始")

    # --------------------------------------------------
    # 2. 列边界检测（保持不变）
    # --------------------------------------------------
    print("\n➡️ 检测列边界（跳过空开头后，最多 10 列）...")
    header_col_end = first_valid_col

    scan_cols = min(MAX_SCAN_COLS, effective_cols)
    if scan_cols >= 1:
        col_consistencies = []
        for j in range(scan_cols):
            actual_j = first_valid_col + j
            col_vals = []
            for di in range(min(5, effective_rows)):
                actual_i = first_valid_row + di
                if actual_j < len(grid[actual_i]):
                    val = safe_str(grid[actual_i][actual_j])
                    if val:
                        col_vals.append(val)
            consistency = compute_column_consistency(col_vals)
            col_consistencies.append(consistency)
            print(f"  有效列 {j} (实际列 {actual_j}): 一致性 = {consistency:.3f} (样本={len(col_vals)})")

        found = False
        for j in range(1, len(col_consistencies)):
            delta = col_consistencies[j] - col_consistencies[j - 1]
            print(f"    Δ(有效列{j}) = {delta:.3f}")
            if delta > DELTA_THRESHOLD:
                header_col_end = first_valid_col + j
                print(f"✅ 列边界确定：数据从实际列 {header_col_end} 开始")
                found = True
                break

        if not found:
            if col_consistencies and col_consistencies[0] > 0.6:
                print("ℹ️ 第一个有效列已具高一致性，无列表头")
                header_col_end = first_valid_col
            else:
                for j, cons in enumerate(col_consistencies):
                    if cons > 0.7:
                        header_col_end = first_valid_col + j
                        print(f"✅ 基于绝对一致性，数据从实际列 {header_col_end} 开始")
                        found = True
                        break
                if not found:
                    print("ℹ️ 未明确检测到数据列，假设从 first_valid_col 开始")
                    header_col_end = first_valid_col
    else:
        print("ℹ️ 有效列数不足，数据从 first_valid_col 开始")

    # --------------------------------------------------
    # 3. 打印边界值（保持不变）
    # --------------------------------------------------
    print("\n📋 边界实际内容:")

    if header_row_end > first_valid_row:
        r = header_row_end - 1
        content = [str(c) if c is not None else '' for c in grid[r][first_valid_col:first_valid_col+10]]
        print(f"  📄 最后一行表头 (行{r}): {content}")
    else:
        print("  📄 无行表头")

    if header_col_end > first_valid_col:
        c = header_col_end - 1
        content = []
        for i in range(first_valid_row, min(first_valid_row + 10, rows)):
            if c < len(grid[i]):
                content.append(str(grid[i][c]) if grid[i][c] is not None else '')
            else:
                content.append('')
        print(f"  📄 最后一列表头 (列{c}): {content}")
    else:
        print("  📄 无列表头")

    if header_row_end < rows and header_col_end < cols:
        val = grid[header_row_end][header_col_end]
        print(f"  🔢 数据起始单元格 ({header_row_end}, {header_col_end}): '{val}'")

    print("\n" + "="*50)
    print("🎯 最终检测结果（已跳过空开头）:")
    print(f"   - 前导空白：行 0~{first_valid_row-1}, 列 0~{first_valid_col-1}")
    print(f"   - 行表头结束于：行 {header_row_end - 1 if header_row_end > first_valid_row else '无'}")
    print(f"   - 列表头结束于：列 {header_col_end - 1 if header_col_end > first_valid_col else '无'}")
    print(f"   - 数据区域起始：(行={header_row_end}, 列={header_col_end})")
    print("="*50)
    return header_row_end, header_col_end


# ==============================
# 使用示例
# ==============================
# if __name__ == "__main__":
#     FILE_PATH = "unmerged_filled_table.xlsx"
#     grid = load_excel_as_grid(FILE_PATH)
#     try:
#         row_end, col_end = detect_table_boundary_skip_empty_start(grid)
#         print(f"\n📌 最终边界：数据起始于 (行={row_end}, 列={col_end})")
#     except Exception as e:
#         print(f"❌ 错误：{e}")
