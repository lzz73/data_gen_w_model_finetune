from .judge_col_value import detect_table_boundary_skip_empty_start, set_embedding_config
from .split_excel import read_excel_and_split, split_table_to_markdown
from .merge_cell import load_excel_as_grid
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def excel_parse(file_path, embedding_config=None, output_format='markdown'):
    """
    解析 Excel 文件

    Args:
        file_path: Excel 文件路径
        embedding_config: 可选，embedding 配置字典，包含 provider, api_key, base_url, model
        output_format: 输出格式 ('markdown' 或 'structured')
            - 'markdown': 输出 Markdown 表格格式，每个工作表作为一个完整的表格
            - 'structured': 输出结构化文本（每行一个切片），适合 recursive 等分割器
    """
    # 只有当 embedding_config 存在且 api_key 非空时才设置
    if embedding_config and embedding_config.get("api_key"):
        set_embedding_config(
            provider=embedding_config.get("provider"),
            api_key=embedding_config.get("api_key"),
            base_url=embedding_config.get("base_url"),
            model=embedding_config.get("model")
        )
        print(f"[DEBUG] Excel parse using provided embedding config")
    else:
        print(f"[DEBUG] No embedding config provided or api_key is empty, will use environment variables")

    # Markdown 格式输出：每个工作表作为一个完整的 Markdown 表格
    if output_format == 'markdown':
        return convert_excel_to_markdown_by_sheet(file_path)

    # 结构化文本输出：每行一个切片
    SHEET_NAME = None
    print("正在对合并单元格进行切分...")
    grid = load_excel_as_grid(file_path)
    print("切分完成，正在检测表格数据边界...")

    # 找到表格边界：有 embedding 配置时使用智能检测，否则使用默认边界（第1行表头）
    if embedding_config and embedding_config.get("api_key"):
        row_end, col_end = detect_table_boundary_skip_empty_start(grid)
        print(f"\n📌 Embedding 边界检测结果：数据起始于 (行={row_end}, 列={col_end})")
    else:
        # 没有 embedding 配置，默认第1行是表头，数据从第2行开始
        row_end = 1
        col_end = 0
        print(f"\n📌 无 Embedding 配置，使用默认边界：数据起始于 (行={row_end}, 列={col_end})")
    # 将表格转化为结构化文本
    texts = read_excel_and_split(
        grid=grid,
        header_row_end=row_end,   # 例如：数据从第 2 行开始（0-indexed）
        header_col_end=col_end,   # 例如：数据从第 1 列开始

    )

    # 打印结果
    for i, text in enumerate(texts[:10]):
        print(f"{i+1}. {text}")
    text = "\n".join(texts)
    return text


def convert_excel_to_markdown_by_sheet(file_path: str) -> list:
    """
    将 Excel 文件按工作表转换为 Markdown 表格
    每个工作表作为一个完整的切片返回（不切分）

    Args:
        file_path: Excel 文件路径

    Returns:
        list: 切片列表，每个元素是一个完整的工作表 Markdown 内容
    """
    wb = load_workbook(file_path, read_only=False, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 获取表格尺寸
        max_row = ws.max_row
        max_col = ws.max_column

        # 处理合并单元格
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col_merge, max_row_merge = range_boundaries(str(merged_range))
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row_merge + 1):
                for col in range(min_col, max_col_merge + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is None:
                        cell.value = top_left_value

        # 构建网格 - 读取所有行和列
        grid = []
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value if cell.value is not None else "")
            grid.append(row_data)

        # 直接转换成 Markdown，整个工作表作为一个表格
        markdown_lines = []

        # 表头行
        header_cells = [str(v) if v is not None else "" for v in grid[0]]
        markdown_lines.append("| " + " | ".join(header_cells) + " |")

        # 分隔线
        markdown_lines.append("| " + " | ".join(["---"] * len(header_cells)) + " |")

        # 数据行（从第 2 行开始）
        for row in grid[1:]:
            row_cells = [str(v) if v is not None else "" for v in row]
            markdown_lines.append("| " + " | ".join(row_cells) + " |")

        markdown_table = "\n".join(markdown_lines)

        # 添加工作表名称
        sheet_content = f"### {sheet_name}\n\n{markdown_table}"
        sheets.append(sheet_content)

    wb.close()

    # 返回切片列表，每个工作表是一个完整的切片
    return sheets

# FILE_PATH = "complex_merged_excel.xlsx"
# excel_parse(file_path=FILE_PATH)
