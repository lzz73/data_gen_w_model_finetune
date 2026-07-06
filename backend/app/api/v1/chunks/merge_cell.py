from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from typing import List, Any, Optional
import os

def load_excel_as_grid(file_path: str, output_path: Optional[str] = None) -> List[List[Any]]:
    """
    加载 Excel 为网格，处理合并单元格。

    合并单元格处理逻辑：
    1. 取消所有合并
    2. 将左上角的值填充到所有空单元格
    3. 使用 cell.value 读取（不是 values_only）
    """
    wb = load_workbook(file_path, read_only=False, data_only=True)
    ws = wb.active

    # 获取表格尺寸
    max_row = ws.max_row
    max_col = ws.max_column

    # 第一步：处理合并单元格
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col_merge, max_row_merge = range_boundaries(str(merged_range))
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row_merge + 1):
            for col in range(min_col, max_col_merge + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    cell.value = top_left_value

    # 第二步：使用 cell.value 读取网格（不是 values_only）
    grid = []
    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(cell.value)
        grid.append(row_data)

    if output_path:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        wb.save(output_path)

    wb.close()
    return grid
